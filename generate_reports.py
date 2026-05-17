"""
generate_reports.py
Queries the SQLite DB and exports a multi-sheet Excel workbook
with formatted KPI tables + a Tableau-ready flat export.
"""
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as xlnums)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import os

DB_PATH      = "/home/claude/analytics_dashboard/data/sales.db"
EXCEL_PATH   = "/home/claude/analytics_dashboard/reports/Sales_Dashboard.xlsx"
TABLEAU_PATH = "/home/claude/analytics_dashboard/reports/Tableau_Ready.csv"

# ── Palette ───────────────────────────────────────────────────────────────────
DARK  = "1A1A2E"
MID   = "16213E"
ACCENT= "0F3460"
GOLD  = "E94560"
WHITE = "FFFFFF"
LIGHT = "F0F4FF"

def qry(conn, sql):
    return pd.read_sql_query(sql, conn)

def style_header(ws, row, ncols, bg=DARK, fg=WHITE, bold=True):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(color=fg, bold=bold, size=11)
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def style_rows(ws, start_row, end_row, ncols):
    fills = [PatternFill("solid", fgColor=LIGHT), PatternFill("solid", fgColor=WHITE)]
    for r in range(start_row, end_row+1):
        fill = fills[(r - start_row) % 2]
        for c in range(1, ncols+1):
            ws.cell(r,c).fill = fill
            ws.cell(r,c).alignment = Alignment(horizontal="center")

def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

def write_sheet(wb, title, df, add_chart=None):
    ws = wb.create_sheet(title)
    # Write header
    for ci, col in enumerate(df.columns, 1):
        ws.cell(1, ci, col.replace("_"," ").title())
    style_header(ws, 1, len(df.columns))
    # Write data
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val)
    style_rows(ws, 2, len(df)+1, len(df.columns))
    autofit(ws)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    if add_chart:
        add_chart(ws, df)
    return ws

def add_revenue_chart(ws, df):
    chart = BarChart()
    chart.type  = "col"
    chart.title = "Monthly Revenue vs Profit"
    chart.y_axis.title = "USD"
    chart.x_axis.title = "Month"
    chart.style = 10
    nrows = len(df) + 1
    rev  = Reference(ws, min_col=4, min_row=1, max_row=nrows)
    prof = Reference(ws, min_col=5, min_row=1, max_row=nrows)
    cats = Reference(ws, min_col=3, min_row=2, max_row=nrows)
    chart.add_data(rev,  titles_from_data=True)
    chart.add_data(prof, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width  = 28
    chart.height = 14
    ws.add_chart(chart, "J2")

def add_region_chart(ws, df):
    chart = BarChart()
    chart.type  = "bar"
    chart.title = "Revenue by Region"
    chart.y_axis.title = "Region"
    nrows = len(df) + 1
    rev  = Reference(ws, min_col=2, min_row=1, max_row=nrows)
    cats = Reference(ws, min_col=1, min_row=2, max_row=nrows)
    chart.add_data(rev, titles_from_data=True)
    chart.set_categories(cats)
    chart.width  = 22
    chart.height = 12
    ws.add_chart(chart, "G2")

def run():
    conn = sqlite3.connect(DB_PATH)

    # ── Overview KPIs ────────────────────────────────────────────────────────
    kpi = qry(conn, """
        SELECT
            COUNT(DISTINCT order_id)   AS Total_Orders,
            ROUND(SUM(revenue),2)      AS Total_Revenue,
            ROUND(SUM(profit),2)       AS Total_Profit,
            ROUND(AVG(margin_pct),2)   AS Avg_Margin_Pct,
            SUM(quantity)              AS Units_Sold,
            ROUND(SUM(revenue)/COUNT(DISTINCT order_id),2) AS Avg_Order_Value
        FROM sales""")

    monthly = qry(conn, """
        SELECT year, month, month_name,
               ROUND(SUM(revenue),2) revenue, ROUND(SUM(profit),2) profit,
               COUNT(*) orders, ROUND(AVG(margin_pct),2) margin_pct
        FROM sales GROUP BY year, month ORDER BY year, month""")

    yoy = qry(conn, """
        SELECT month, month_name,
               ROUND(SUM(CASE WHEN year=2023 THEN revenue ELSE 0 END),2) rev_2023,
               ROUND(SUM(CASE WHEN year=2024 THEN revenue ELSE 0 END),2) rev_2024,
               ROUND((SUM(CASE WHEN year=2024 THEN revenue ELSE 0 END)-
                      SUM(CASE WHEN year=2023 THEN revenue ELSE 0 END))/
                      NULLIF(SUM(CASE WHEN year=2023 THEN revenue ELSE 0 END),0)*100,2) yoy_pct
        FROM sales GROUP BY month, month_name ORDER BY month""")

    region = qry(conn, "SELECT * FROM v_region_kpis ORDER BY total_revenue DESC")
    category = qry(conn, "SELECT * FROM v_category_kpis ORDER BY total_revenue DESC")
    reps   = qry(conn, "SELECT * FROM v_rep_kpis")
    channel= qry(conn, """
        SELECT channel, ROUND(SUM(revenue),2) revenue, ROUND(SUM(profit),2) profit,
               COUNT(*) orders,
               ROUND(SUM(revenue)*100.0/SUM(SUM(revenue)) OVER(),1) revenue_share_pct
        FROM sales GROUP BY channel ORDER BY revenue DESC""")
    quarterly = qry(conn, """
        SELECT year, quarter, ROUND(SUM(revenue),2) revenue,
               ROUND(SUM(profit),2) profit, COUNT(*) orders,
               ROUND(AVG(margin_pct),2) avg_margin
        FROM sales GROUP BY year, quarter ORDER BY year, quarter""")

    conn.close()

    # ── Build workbook ───────────────────────────────────────────────────────
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # Summary page
    ws0 = wb.create_sheet("📊 Overview")
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 28
    ws0.column_dimensions["B"].width = 22

    titles = ["Total Orders","Total Revenue ($)","Total Profit ($)",
              "Avg Margin (%)","Units Sold","Avg Order Value ($)"]
    vals   = kpi.iloc[0].tolist()
    ws0.cell(1,1,"Sales Dashboard — KPI Overview").font = Font(size=16, bold=True, color=DARK)
    ws0.cell(1,1).alignment = Alignment(horizontal="left")
    ws0.merge_cells("A1:B1")
    ws0.row_dimensions[1].height = 30

    for i,(t,v) in enumerate(zip(titles,vals), 3):
        ws0.cell(i,1,t).font = Font(bold=True, color=DARK, size=11)
        c = ws0.cell(i,2,v)
        c.font = Font(color=ACCENT, bold=True, size=12)
        c.alignment = Alignment(horizontal="right")
        ws0.cell(i,1).fill = PatternFill("solid", fgColor=LIGHT)
        ws0.cell(i,2).fill = PatternFill("solid", fgColor=LIGHT)

    # Data sheets
    write_sheet(wb, "📅 Monthly Trend",   monthly,   add_revenue_chart)
    write_sheet(wb, "📆 YoY Comparison",  yoy)
    write_sheet(wb, "🗺 Region KPIs",     region,    add_region_chart)
    write_sheet(wb, "🏷 Category KPIs",   category)
    write_sheet(wb, "👤 Sales Reps",      reps)
    write_sheet(wb, "📡 Channel Mix",     channel)
    write_sheet(wb, "📈 Quarterly",       quarterly)

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"Excel saved → {EXCEL_PATH}")

    # ── Tableau flat export ──────────────────────────────────────────────────
    conn2 = sqlite3.connect(DB_PATH)
    flat  = pd.read_sql_query("SELECT * FROM sales", conn2)
    conn2.close()
    flat.to_csv(TABLEAU_PATH, index=False)
    print(f"Tableau CSV saved → {TABLEAU_PATH}")

if __name__ == "__main__":
    run()
